import openpyxl as xl
from openpyxl.chart import BarChart, Reference
from pathlib import Path
import utils

def process_workbook(filename):

    wb = xl.load_workbook(filename)
    sheet = wb["Sheet1"]

    rowNum = sheet.max_row

    for row in range(2,rowNum+1):
        cell = sheet.cell(row,3)
        corrected_price = cell.value * 0.9
        corrected_cell = sheet.cell(row,4)
        corrected_cell.value = corrected_price

    values = Reference(sheet,min_row=2,max_row=sheet.max_row,min_col=4,max_col=4)
    chart = BarChart()
    chart.add_data(values)

    sheet.add_chart(chart,'e2')

    wb.save(filename)

process_workbook("transactions.xlsx")

